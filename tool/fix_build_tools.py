import os
import sys
import shutil
import subprocess
import json
import openpyxl
from collections import deque
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Set
from google.adk.tools.tool_context import ToolContext

CURRENT_TOOL_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(CURRENT_TOOL_DIR)

PROCESSED_PROJECTS_DIR = os.path.join(ROOT_DIR, "process")
PROCESSED_PROJECTS_FILE = os.path.join(PROCESSED_PROJECTS_DIR, "project_processed.txt")

# Core Tools
def force_clean_git_repo(repo_path: str) -> Dict[str, str]:
    print(f"--- Tool: force_clean_git_repo (v2) called for: {repo_path} ---")

    if not os.path.isdir(os.path.join(repo_path, ".git")):
        return {'status': 'error', 'message': f"Directory '{repo_path}' is not a valid Git repository."}

    original_path = os.getcwd()
    try:
        os.chdir(repo_path)

        # 1. First, switch to the main branch.
        # 2. Force reset to HEAD.
        subprocess.run(["git", "reset", "--hard", "HEAD"], capture_output=True, text=True, check=True)

        # 3. Switch to main/master branch.
        main_branch = "main" if "main" in subprocess.run(["git", "branch", "--list"], capture_output=True, text=True).stdout else "master"
        subprocess.run(["git", "switch", main_branch], capture_output=True, text=True, check=True)

        # 4. Remove all untracked files.
        subprocess.run(["git", "clean", "-fdx"], capture_output=True, text=True, check=True)

        message = f"Successfully force-cleaned the repository '{repo_path}'. All local changes and untracked files have been removed."
        print(message)
        return {'status': 'success', 'message': message}

    except subprocess.CalledProcessError as e:
        message = f"Failed to force-clean repository '{repo_path}': {e.stderr.strip()}"
        print(f"--- ERROR: {message} ---")
        return {'status': 'error', 'message': message}
    except Exception as e:
        message = f"An unknown error occurred while cleaning the repository: {e}"
        print(f"--- ERROR: {message} ---")
        return {'status': 'error', 'message': message}
    finally:
        os.chdir(original_path)


def get_project_paths(project_name: str) -> Dict[str, str]:
    """
    Generates and returns the standard project_config_path and project_source_path based on the project name.
    """
    print(f"--- Tool: get_project_paths called for: {project_name} ---")
    base_path = os.path.abspath(os.path.join(os.path.dirname(__file__)))

    safe_project_name = "".join(c for c in project_name if c.isalnum() or c in ('_', '-')).rstrip()

    config_path = os.path.join(base_path, "oss-fuzz", "projects", safe_project_name)
    source_path = os.path.join(base_path, "process", "project", safe_project_name)

    paths = {
        "project_name": project_name,
        "project_config_path": config_path,
        "project_source_path": source_path,
        "max_depth": 1
    }
    print(f"--- Generated paths: {paths} ---")
    return paths


def save_processed_project(project_name: str) -> Dict[str, str]:
    """
    Appends a processed project name to the project_processed.txt file.
    """
    print(f"--- Tool: save_processed_project called for: {project_name} ---")
    try:
        os.makedirs(PROCESSED_PROJECTS_DIR, exist_ok=True)
        with open(PROCESSED_PROJECTS_FILE, 'a', encoding='utf-8') as f:
            f.write(f"{project_name}\n")
        message = f"Successfully saved '{project_name}' to processed list."
        print(f"--- {message} ---")
        return {"status": "success", "message": message}
    except Exception as e:
        message = f"Failed to save processed project '{project_name}': {e}"
        print(f"--- ERROR: {message} ---")
        return {"status": "error", "message": message}

def update_excel_report(file_path: str, row_index: int, attempted: str, result: str) -> Dict[str, str]:
    """
    [Revised] Updates the "Fix Attempted", "Fix Result", and "Fix Date" columns for a specified row in an .xlsx file.
    """
    print(f"--- Tool: update_excel_report called for file '{file_path}', row {row_index} ---")
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]

        # Dynamically get column indices (Translated headers)
        attempted_col_idx = headers.index("Fix Attempted") + 1
        result_col_idx = headers.index("Fix Result") + 1
        date_col_idx = headers.index("Fix Date") + 1

        sheet.cell(row=row_index, column=attempted_col_idx, value=attempted)
        sheet.cell(row=row_index, column=result_col_idx, value=result)
        sheet.cell(row=row_index, column=date_col_idx, value=datetime.now().strftime('%Y-%m-%d'))

        workbook.save(file_path)
        message = f"Successfully updated row {row_index} in '{file_path}' with result: '{result}'."
        print(message)
        return {'status': 'success', 'message': message}
    except Exception as e:
        message = f"Failed to update Excel file: {e}"
        print(f"--- ERROR: {message} ---")
        return {'status': 'error', 'message': message}


def read_projects_from_excel(file_path: str) -> Dict:
    """
    [Revised] Reads project information from the specified .xlsx file.
    """
    print(f"--- Tool: read_projects_from_excel called for: {file_path} ---")
    if not os.path.exists(file_path):
        return {'status': 'error', 'message': f"Excel file not found at '{file_path}'."}

    projects_to_run = []
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]

        # Verify required headers (Translated)
        required_headers = ["Project Name", "Reproduce OSS-Fuzz SHA", "Error Consistency", "Fix Attempted"]
        if not all(h in headers for h in required_headers):
             return {'status': 'error', 'message': f"Excel file is missing one of the required columns: {required_headers}"}

        name_idx = headers.index("Project Name")
        sha_idx = headers.index("Reproduce OSS-Fuzz SHA")
        consistent_idx = headers.index("Error Consistency")
        attempted_idx = headers.index("Fix Attempted")

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # "Yes" check
            if row[consistent_idx] == "Yes" and row[attempted_idx] != "Yes":
                project_info = {
                    "project_name": row[name_idx],
                    "sha": str(row[sha_idx]),
                    "row_index": row_index
                }
                projects_to_run.append(project_info)

        print(f"--- Found {len(projects_to_run)} new projects to process. ---")
        return {'status': 'success', 'projects': projects_to_run}
    except Exception as e:
        return {'status': 'error', 'message': f"Failed to read or parse Excel file: {e}"}


def run_command(command: str) -> Dict[str, str]:
    """
    Executes a shell command and returns its output.
    """
    print(f"--- Tool: run_command called with: '{command}' ---")
    try:
        result = subprocess.run(
            command,
            shell=True,
            capture_output=True,
            text=True,
            check=True,
            encoding='utf-8'
        )
        output = f"STDOUT:\n{result.stdout}\nSTDERR:\n{result.stderr}"
        return {"status": "success", "output": output}
    except subprocess.CalledProcessError as e:
        output = f"Error executing command.\nReturn Code: {e.returncode}\nSTDOUT:\n{e.stdout}\nSTDERR:\n{e.stderr}"
        return {"status": "error", "message": output}
    except Exception as e:
        return {"status": "error", "message": f"An unexpected error occurred: {e}"}

def truncate_prompt_file(file_path: str, max_lines: int = 2000) -> Dict[str, str]:
    """
    Reads a file, and if it exceeds max_lines, truncates it.
    """
    print(f"--- Tool: truncate_prompt_file called for: {file_path} ---")
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        if len(lines) <= max_lines:
            message = "File is within line limits, no truncation needed."
            print(f"--- {message} ---")
            return {"status": "success", "message": message}

        head_count = max_lines // 4
        tail_count = max_lines - head_count

        truncated_content = "".join(lines[:head_count])
        truncated_content += "\n\n... (Content truncated due to context length limit) ...\n\n"
        truncated_content += "".join(lines[-tail_count:])

        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(truncated_content)

        message = f"File '{file_path}' was truncated to approximately {max_lines} lines."
        print(f"--- {message} ---")
        return {"status": "success", "message": message}
    except Exception as e:
        message = f"Failed to truncate file '{file_path}': {e}"
        print(f"--- ERROR: {message} ---")
        return {"status": "error", "message": message}

def archive_fixed_project(project_name: str, project_config_path: str) -> Dict[str, str]:
    """
    Archives the configuration directory of a successfully fixed project.
    """
    print(f"--- Tool: archive_fixed_project called for: {project_name} ---")
    try:
        base_success_dir = "success-fix-project"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_project_name = "".join(c for c in project_name if c.isalnum() or c in ('_', '-')).rstrip()

        destination_dir = os.path.join(base_success_dir, f"{safe_project_name}_{timestamp}")

        if not os.path.isdir(project_config_path):
            return {"status": "error", "message": f"Source config path does not exist: {project_config_path}"}

        shutil.copytree(project_config_path, destination_dir)

        message = f"Successfully archived config files for '{project_name}' to '{destination_dir}'."
        print(f"--- {message} ---")
        return {"status": "success", "message": message}
    except Exception as e:
        message = f"Failed to archive project '{project_name}': {e}"
        print(f"--- ERROR: {message} ---")
        return {"status": "error", "message": message}


def download_github_repo(project_name: str, target_dir: str) -> Dict[str, str]:
    """
    Searches for a project on GitHub and clones it.
    """
    print(f"--- Tool: download_github_repo called for '{project_name}' into '{target_dir}' ---")

    if os.path.isdir(target_dir):
        if project_name == "oss-fuzz":
             print(f"--- Directory '{target_dir}' already exists. Pulling latest changes. ---")
             try:
                 subprocess.run(["git", "pull"], cwd=target_dir, check=True, capture_output=True)
             except Exception as e:
                 print(f"Warning: Failed to pull latest changes for oss-fuzz: {e}")
        else:
            print(f"--- Directory '{target_dir}' already exists. Skipping download. ---")
        return {'status': 'success', 'path': target_dir}

    os.makedirs(os.path.dirname(target_dir), exist_ok=True)

    try:
        if project_name == "oss-fuzz":
            repo_full_name = "google/oss-fuzz"
        else:
            search_command = ["gh", "search", "repos", project_name, "--sort", "stars", "--order", "desc", "--limit", "1", "--json", "fullName"]
            result = subprocess.run(search_command, capture_output=True, text=True, check=True, encoding='utf-8')
            parsed_output = json.loads(result.stdout.strip())
            if isinstance(parsed_output, list) and parsed_output:
                repo_full_name = parsed_output[0]['fullName']
            else: raise ValueError("gh search command returned unexpected JSON.")

        repo_url = f"https://github.com/{repo_full_name}.git"
        print(f"--- Found repository URL: {repo_url} ---")
    except Exception as e:
        message = f"ERROR: 'gh' CLI search failed for '{project_name}'. Details: {e}"
        return {'status': 'error', 'message': message}

    clone_command = ["git", "clone", repo_url, target_dir]
    if project_name != "oss-fuzz":
        clone_command.insert(2, "--depth=1")

    try:
        subprocess.run(clone_command, check=True, capture_output=True, text=True)
        message = f"Successfully cloned '{project_name}' to '{target_dir}'."
        return {'status': 'success', 'path': target_dir, 'message': message}
    except subprocess.CalledProcessError as e:
        message = f"Git clone failed for '{project_name}': {e.stderr}"
        return {'status': 'error', 'message': message}


def find_sha_for_timestamp(commits_file_path: str, error_date: str) -> Dict[str, str]:
    """
    Finds the most suitable commit SHA for a given date.
    """
    print(f"--- Tool: find_sha_for_timestamp called for date: {error_date} ---")
    try:
        target_date = datetime.strptime(error_date, '%Y.%m.%d').date()
    except ValueError:
        return {'status': 'error', 'message': f"Invalid target date format: '{error_date}'. Expected 'YYYY.MM.DD'."}

    todays_commits: List[Tuple[datetime, str]] = []
    past_commits: List[Tuple[datetime, str]] = []

    try:
        with open(commits_file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if line.startswith("Time: ") and i + 1 < len(lines) and lines[i+1].strip().startswith("- SHA: "):
                try:
                    timestamp_str = line.replace("Time: ", "")
                    commit_datetime = datetime.strptime(timestamp_str, '%Y.%m.%d %H:%M')
                    sha = lines[i+1].strip().replace("- SHA: ", "")
                    commit_date = commit_datetime.date()
                    if commit_date == target_date:
                        todays_commits.append((commit_datetime, sha))
                    elif commit_date < target_date:
                        past_commits.append((commit_datetime, sha))
                except (ValueError, IndexError):
                    pass
            i += 1
    except FileNotFoundError:
        return {'status': 'error', 'message': f"Commits file not found at: {commits_file_path}"}
    except Exception as e:
        return {'status': 'error', 'message': f"An unexpected error occurred: {e}"}

    if todays_commits:
        earliest_today = min(todays_commits)
        found_sha = earliest_today[1]
        return {'status': 'success', 'sha': found_sha}
    elif past_commits:
        latest_in_past = max(past_commits)
        found_sha = latest_in_past[1]
        return {'status': 'success', 'sha': found_sha}
    else:
        return {'status': 'error', 'message': f"No suitable SHA found on or before the date {error_date}."}


def checkout_oss_fuzz_commit(sha: str) -> Dict[str, str]:
    """
    Executes a git checkout command in the fixed oss-fuzz directory.
    """
    oss_fuzz_path = os.path.join(ROOT_DIR, "oss-fuzz")
    print(f"--- Tool: checkout_oss_fuzz_commit called for SHA: {sha} in '{oss_fuzz_path}' ---")

    if not os.path.isdir(os.path.join(oss_fuzz_path, ".git")):
        return {'status': 'error', 'message': f"The directory '{oss_fuzz_path}' is not a git repository."}

    original_path = os.getcwd()
    try:
        os.chdir(oss_fuzz_path)
        main_branch = "main" if "main" in subprocess.run(["git", "branch"], capture_output=True, text=True).stdout else "master"
        subprocess.run(["git", "switch", main_branch], capture_output=True, text=True)

        command = ["git", "checkout", sha]
        result = subprocess.run(command, capture_output=True, text=True, encoding='utf-8')

        if result.returncode == 0:
            return {'status': 'success', 'message': f"Successfully checked out SHA {sha}."}
        else:
            return {'status': 'error', 'message': f"Git command failed: {result.stderr.strip()}"}
    except Exception as e:
        return {'status': 'error', 'message': f"An unexpected error occurred during checkout: {e}"}
    finally:
        os.chdir(original_path)

# File Operations and Fuzzing Tools

def apply_patch(solution_file_path: str) -> dict:
    """
    Reads a specially formatted solution file and applies the code replacement solution within it.
    """
    print(f"--- Tool: apply_patch (New Version) called for solution file: {solution_file_path} ---")
    try:
        with open(solution_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        file_part = content.split('---=== FILE ===---')[1].strip()
        original_part = file_part.split('---=== ORIGINAL ===---')[1].strip()
        replacement_part = original_part.split('---=== REPLACEMENT ===---')[1].strip()
        file_path = file_part.split('---=== ORIGINAL ===---')[0].strip()
        original_block = original_part.split('---=== REPLACEMENT ===---')[0].strip()
        replacement_block = replacement_part
        if not file_path or not original_block:
            return {"status": "error", "message": "Solution file format is incorrect. Could not parse FILE path or ORIGINAL block."}
        if not os.path.exists(file_path):
            return {"status": "error", "message": f"Target file does not exist: {file_path}"}
        with open(file_path, 'r', encoding='utf-8') as f:
            original_content = f.read()
        if original_block not in original_content:
            return {"status": "error", "message": "The ORIGINAL code block was not found in the target file. The file may have already been modified or the block is incorrect."}
        new_content = original_content.replace(original_block, replacement_block, 1)
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        success_message = f"Successfully applied code fix to '{file_path}'."
        print(success_message)
        return {"status": "success", "message": success_message}
    except IndexError:
        error_message = "Failed to parse solution file. Make sure it contains FILE, ORIGINAL, and REPLACEMENT separators."
        print(error_message)
        return {"status": "error", "message": error_message}
    except Exception as e:
        error_message = f"An error occurred while applying the code fix: {str(e)}"
        print(error_message)
        return {"status": "error", "message": error_message}

def save_file_tree(directory_path: str, output_file: Optional[str] = None) -> dict:
    """
    Gets the file tree structure of a specified directory path and saves it to a file.
    """
    print(f"--- Tool: save_file_tree called for path: {directory_path} ---")
    if not os.path.isdir(directory_path):
        error_message = f"Error: The provided path '{directory_path}' is not a valid directory."
        print(error_message)
        return {"status": "error", "message": error_message}
    if output_file is None:
        output_dir = "generated_prompt_file"
        final_output_path = os.path.join(output_dir, "file_tree.txt")
    else:
        final_output_path = output_file
    output_dir = os.path.dirname(final_output_path)
    try:
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        tree_lines = []
        def _build_tree_recursive(path, prefix=""):
            entries = sorted([e for e in os.listdir(path) if not e.startswith('.')])
            pointers = ["├── "] * (len(entries) - 1) + ["└── "]
            for pointer, entry in zip(pointers, entries):
                full_path = os.path.join(path, entry)
                if os.path.isdir(full_path):
                    tree_lines.append(f"{prefix}{pointer}📁 {entry}")
                    extension = "│   " if pointer == "├── " else "    "
                    _build_tree_recursive(full_path, prefix + extension)
                else:
                    tree_lines.append(f"{prefix}{pointer}📄 {entry}")
        tree_lines.insert(0, f"📁 {os.path.basename(os.path.abspath(directory_path))}")
        _build_tree_recursive(directory_path, prefix="")
        with open(final_output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(tree_lines))
        success_message = f"File tree has been successfully generated and saved to '{final_output_path}'."
        print(success_message)
        return {"status": "success", "message": success_message}
    except Exception as e:
        error_message = f"An error occurred while generating or saving the file tree: {str(e)}"
        print(error_message)
        return {"status": "error", "message": error_message}

def save_file_tree_shallow(directory_path: str, max_depth: int, output_file: Optional[str] = None) -> dict:
    """
    Gets the top N levels of the file tree structure.
    """
    print(f"--- Tool: save_file_tree_shallow called for path: {directory_path} with max_depth: {max_depth} ---")
    if not os.path.isdir(directory_path):
        error_message = f"Error: The provided path '{directory_path}' is not a valid directory."
        print(error_message)
        return {"status": "error", "message": error_message}
    if output_file is None:
        output_dir = "generated_prompt_file"
        final_output_path = os.path.join(output_dir, "file_tree.txt")
    else:
        final_output_path = output_file
    output_dir = os.path.dirname(final_output_path)
    try:
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        tree_lines = []
        def _build_tree_recursive(path, prefix="", depth=0):
            if depth >= max_depth:
                return
            try:
                entries = sorted([e for e in os.listdir(path) if not e.startswith('.')])
            except OSError:
                entries = []
            pointers = ["├── "] * (len(entries) - 1) + ["└── "]
            for pointer, entry in zip(pointers, entries):
                full_path = os.path.join(path, entry)
                if os.path.isdir(full_path):
                    tree_lines.append(f"{prefix}{pointer}📁 {entry}")
                    extension = "│   " if pointer == "├── " else "    "
                    _build_tree_recursive(full_path, prefix + extension, depth + 1)
                else:
                    tree_lines.append(f"{prefix}{pointer}📄 {entry}")
        tree_lines.insert(0, f"📁 {os.path.basename(os.path.abspath(directory_path))}")
        _build_tree_recursive(directory_path, prefix="")
        with open(final_output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(tree_lines))
        success_message = f"The top {max_depth} levels of the file tree have been successfully generated and saved to '{final_output_path}'."
        print(success_message)
        return {"status": "success", "message": success_message}
    except Exception as e:
        error_message = f"An error occurred while generating or saving the shallow file tree: {str(e)}"
        print(error_message)
        return {"status": "error", "message": error_message}

def find_and_append_file_details(directory_path: str, search_keyword: str, output_file: Optional[str] = None) -> dict:
    """
    Finds a file or directory by its name or partial path and appends its detailed structure to a file.
    """
    print(f"--- Tool: find_and_append_file_details called for path: {directory_path} with keyword: '{search_keyword}' ---")
    if not os.path.isdir(directory_path):
        error_message = f"Error: The provided path '{directory_path}' is not a valid directory."
        print(error_message)
        return {"status": "error", "message": error_message}
    if output_file is None:
        output_dir = "generated_prompt_file"
        final_output_path = os.path.join(output_dir, "file_tree.txt")
    else:
        final_output_path = output_file
    output_dir = os.path.dirname(final_output_path)
    try:
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        found_paths = []
        for root, dirs, files in os.walk(directory_path):
            all_entries = dirs + files
            for entry in all_entries:
                full_path = os.path.join(root, entry)
                if search_keyword in full_path:
                    found_paths.append(full_path)
        found_paths = sorted(list(set(found_paths)))
        if not found_paths:
            message = f"No file or directory matching '{search_keyword}' was found in '{directory_path}'."
            print(message)
            with open(final_output_path, "a", encoding="utf-8") as f:
                f.write(f"\n\n--- Detailed query result for '{search_keyword}' ---\n")
                f.write(message)
            return {"status": "success", "message": message}
        details_to_append = [f"\n\n--- Detailed query result for '{search_keyword}' ---"]
        for path in found_paths:
            relative_path = os.path.relpath(path, directory_path)
            details_to_append.append(f"\n# Matched path: {relative_path}")
            if os.path.isdir(path):
                def _build_tree_recursive(sub_path, prefix=""):
                    try:
                        entries = sorted([e for e in os.listdir(sub_path) if not e.startswith('.')])
                    except OSError:
                        entries = []
                    pointers = ["├── "] * (len(entries) - 1) + ["└── "]
                    for pointer, entry in zip(pointers, entries):
                        details_to_append.append(f"{prefix}{pointer}{'📁' if os.path.isdir(os.path.join(sub_path, entry)) else '📄'} {entry}")
                _build_tree_recursive(path)
            else:
                details_to_append.append(f"📄 {os.path.basename(path)}")
        with open(final_output_path, "a", encoding="utf-8") as f:
            f.write("\n".join(details_to_append))
        success_message = f"Detailed search results for '{search_keyword}' have been appended to '{final_output_path}'."
        print(success_message)
        return {"status": "success", "message": success_message}
    except Exception as e:
        error_message = f"An error occurred while finding and appending file details: {str(e)}"
        print(error_message)
        return {"status": "error", "message": error_message}

def read_file_content(file_path: str) -> dict:
    """
    Reads and returns the content of a specified text file.
    """
    print(f"--- Tool: read_file_content called for path: {file_path} ---")
    MAX_FILE_SIZE = 1024 * 1024
    if not os.path.exists(file_path):
        message = f"Error: File '{file_path}' does not exist."
        print(message)
        return {"status": "error", "message": message}
    if not os.path.isfile(file_path):
        message = f"Error: Path '{file_path}' is a directory, not a file."
        print(message)
        return {"status": "error", "message": message}
    if os.path.getsize(file_path) > MAX_FILE_SIZE:
        message = f"Error: File '{file_path}' is too large to process."
        print(message)
        return {"status": "error", "message": message}
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
        success_message = f"Content of file '{file_path}' has been successfully read into memory."
        print(success_message)
        return {"status": "success", "message": success_message, "content": content}
    except Exception as e:
        message = f"An error occurred while reading file '{file_path}': {str(e)}"
        print(message)
        return {"status": "error", "message": message}

def create_or_update_file(file_path: str, content: str) -> dict:
    """
    Creates a new file and writes content to it, or overwrites an existing file.
    """
    print(f"--- Tool: create_or_update_file called for path: {file_path} ---")
    try:
        directory = os.path.dirname(file_path)
        if directory:
            os.makedirs(directory, exist_ok=True)
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(content)
        message = f"File '{file_path}' has been successfully created/updated."
        print(message)
        return {"status": "success", "message": message}
    except Exception as e:
        message = f"An error occurred while creating or updating file '{file_path}': {str(e)}"
        print(message)
        return {"status": "error", "message": message}

def append_file_to_file(source_path: str, destination_path: str) -> dict:
    """
    Reads the entire content of a source file and appends it to the end of a destination file.
    """
    print(f"--- Tool: append_file_to_file called. Source: '{source_path}', Destination: '{destination_path}' ---")
    if not os.path.isfile(source_path):
        return {"status": "error", "message": f"Error: Source file '{source_path}' does not exist or is not a valid file."}
    if os.path.isdir(destination_path):
        return {"status": "error", "message": f"Error: Destination path '{destination_path}' is a directory and cannot be an append target."}
    if os.path.abspath(source_path) == os.path.abspath(destination_path):
        return {"status": "error", "message": "Error: Source and destination files cannot be the same."}
    try:
        with open(source_path, "r", encoding="utf-8") as f_source:
            content_to_append = f_source.read()
        dest_directory = os.path.dirname(destination_path)
        if dest_directory:
            os.makedirs(dest_directory, exist_ok=True)
        with open(destination_path, "a", encoding="utf-8") as f_dest:
            f_dest.write(content_to_append)
        return {"status": "success", "message": f"Successfully appended the content of '{source_path}' to '{destination_path}'."}
    except Exception as e:
        return {"status": "error", "message": f"An unknown error occurred while appending the file: {str(e)}"}

def append_string_to_file(file_path: str, content: str) -> dict:
    """
    Appends a string of content to the end of a specified file.
    """
    print(f"--- Tool: append_string_to_file called for path: {file_path} ---")
    try:
        directory = os.path.dirname(file_path)
        if directory:
            os.makedirs(directory, exist_ok=True)
        with open(file_path, "a", encoding="utf-8") as f:
            f.write(content)
        return {"status": "success", "message": f"Content successfully appended to file '{file_path}'."}
    except Exception as e:
        return {"status": "error", "message": f"An error occurred while appending content to file '{file_path}': {str(e)}"}

def delete_file(file_path: str) -> dict:
    """
    Deletes a specified file.
    """
    print(f"--- Tool: delete_file called for path: {file_path} ---")
    if not os.path.exists(file_path):
        message = f"Error: File '{file_path}' does not exist and cannot be deleted."
        print(message)
        return {"status": "error", "message": message}
    try:
        os.remove(file_path)
        message = f"File '{file_path}' has been successfully deleted."
        print(message)
        return {"status": "success", "message": message}
    except Exception as e:
        message = f"An error occurred while deleting file '{file_path}': {str(e)}"
        print(message)
        return {"status": "error", "message": message}

def prompt_generate_tool(project_main_folder_path: str, max_depth: int, config_folder_path: str) -> dict:
    """
    Automatically collects various fuzzing context information and integrates it into a single prompt file.
    """
    print("--- Workflow Tool: prompt_generate_tool started ---")
    PROMPT_DIR = "generated_prompt_file"
    PROMPT_FILE_PATH = os.path.join(PROMPT_DIR, "prompt.txt")
    FILE_TREE_PATH = os.path.join(PROMPT_DIR, "file_tree.txt")
    FUZZ_LOG_PATH = "fuzz_build_log_file/fuzz_build_log.txt"
    print(f"Step 0: Discovering configuration files in '{config_folder_path}'...")
    if not os.path.isdir(config_folder_path):
        return {"status": "error", "message": f"Error: The provided config path '{config_folder_path}' is not a valid directory."}
    try:
        all_config_files = [
            os.path.join(config_folder_path, f)
            for f in sorted(os.listdir(config_folder_path))
            if os.path.isfile(os.path.join(config_folder_path, f))
        ]
        if not all_config_files:
            print(f"Warning: No files were found in the directory '{config_folder_path}'.")
    except Exception as e:
        return {"status": "error", "message": f"An error occurred while scanning the config directory: {str(e)}"}
    print("Step 1: Generating and writing the introductory prompt...")
    project_name = os.path.basename(os.path.abspath(project_main_folder_path))
    config_file_names = [os.path.basename(f) for f in all_config_files]
    config_files_str = ", ".join(config_file_names) if config_file_names else "(None)"
    introductory_prompt = f"""
You are a premier expert in software testing, specializing in solving fuzzing compilation and build issues. These problems are often caused by mismatches between fuzzing configuration files and the project's source files. I will provide you with error logs from the oss-fuzz build process for different projects. Based on the error messages, configuration files, and other information, you are to provide targeted solutions. Strive to avoid altering files unrelated to the problem, with the ultimate goal of enabling the project to compile and build successfully.
Next, the {config_files_str}, file tree, and error log for {project_name} will be provided. Please read and analyze the file tree and the given information, identify which files might be causing the problem—whether they are core fuzz testing build files like Dockerfile and build.sh, or files within the {project_name} project itself—and attempt to propose a solution.
"""
    os.makedirs(PROMPT_DIR, exist_ok=True)
    with open(PROMPT_FILE_PATH, "w", encoding="utf-8") as f:
        f.write(introductory_prompt)
    print("Step 2: Appending configuration files...")
    with open(PROMPT_FILE_PATH, "a", encoding="utf-8") as f:
        f.write("\n\n--- Configuration Files ---\n")
    for config_file in all_config_files:
        with open(PROMPT_FILE_PATH, "a", encoding="utf-8") as f:
            f.write(f"\n### Content from: {os.path.basename(config_file)} ###\n")
        print(f"  - Appending '{config_file}'...")
        try:
            with open(config_file, "r", encoding="utf-8") as source_f, open(PROMPT_FILE_PATH, "a", encoding="utf-8") as dest_f:
                dest_f.write(source_f.read())
        except Exception as e:
            print(f"    Warning: Failed to append '{config_file}': {e}. Skipping.")
    print(f"Step 3: Generating shallow project file tree (max_depth='{max_depth}')...")
    result = save_file_tree_shallow(
        directory_path=project_main_folder_path,
        max_depth=max_depth,
        output_file=FILE_TREE_PATH
    )
    if result["status"] == "error":
        return result
    print("Step 4: Appending file tree to prompt file...")
    with open(PROMPT_FILE_PATH, "a", encoding="utf-8") as f:
        f.write("\n\n--- Project File Tree (Shallow View) ---\n")
    try:
        with open(FILE_TREE_PATH, "r", encoding="utf-8") as source_f, open(PROMPT_FILE_PATH, "a", encoding="utf-8") as dest_f:
            dest_f.write(source_f.read())
    except Exception as e:
        return {"status": "error", "message": f"Failed to append file tree: {e}"}
    print("Step 5: Checking for and appending fuzz build log...")
    if os.path.isfile(FUZZ_LOG_PATH) and os.path.getsize(FUZZ_LOG_PATH) > 0:
        print(f"  - Found fuzz log at '{FUZZ_LOG_PATH}'. Appending...")
        with open(PROMPT_FILE_PATH, "a", encoding="utf-8") as f:
            f.write("\n\n--- Fuzz Build Log ---\n")
        try:
            with open(FUZZ_LOG_PATH, "r", encoding="utf-8") as source_f, open(PROMPT_FILE_PATH, "a", encoding="utf-8") as dest_f:
                dest_f.write(source_f.read())
        except Exception as e:
            print(f"    Warning: Failed to append fuzz log: {e}.")
    else:
        print("  - Fuzz log not found or is empty. Skipping.")
    final_message = (
        f"Prompt generation workflow completed successfully. Initial context information has been consolidated into '{PROMPT_FILE_PATH}'. "
        f"This includes the project's file structure up to '{max_depth}' levels deep. Please analyze the existing information. If you need to delve deeper into a specific directory, "
        f"use the 'find_and_append_file_details' tool for a precise search."
    )
    print(f"--- Workflow Tool: prompt_generate_tool finished successfully ---")
    return {"status": "success", "message": final_message}


def run_fuzz_build_streaming(
    project_name: str,
    oss_fuzz_path: str,
    sanitizer: str,
    engine: str,
    architecture: str
) -> dict:
    """
    Executes a predefined fuzzing build command and streams its output in real-time.
    """
    print(f"--- Tool: run_fuzz_build_streaming called for project: {project_name} ---")

    target_oss_fuzz_path = os.path.join(ROOT_DIR, "oss-fuzz")

    if not os.path.exists(oss_fuzz_path) or not os.path.isabs(oss_fuzz_path):
        print(f"--- Path Correction: Redirecting '{oss_fuzz_path}' to absolute path '{target_oss_fuzz_path}' ---")
        oss_fuzz_path = target_oss_fuzz_path

    helper_script_path = os.path.join(oss_fuzz_path, "infra", "helper.py")
    if not os.path.exists(helper_script_path):
        error_msg = f"Critical Error: 'infra/helper.py' not found at expected path: {helper_script_path}"
        print(f"--- {error_msg} ---")
        return {"status": "error", "message": error_msg}

    LOG_DIR = "fuzz_build_log_file"
    LOG_FILE_PATH = os.path.join(LOG_DIR, "fuzz_build_log.txt")

    try:
        command = [
            "python3.10", helper_script_path, "build_fuzzers",
            "--sanitizer", sanitizer,
            "--engine", engine,
            "--architecture", architecture,
            project_name
        ]

        print(f"--- Executing command: {' '.join(command)} ---")
        print(f"--- Working Directory: {oss_fuzz_path} ---")
        print("--- Fuzzing process started. Real-time output will be displayed below: ---")

        process = subprocess.Popen(
            command,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
            cwd=oss_fuzz_path,
            encoding='utf-8'
        )

        log_buffer = deque(maxlen=280)
        for line in process.stdout:
            print(line, end='', flush=True)
            log_buffer.append(line)

        process.wait()
        return_code = process.returncode

        print("\n--- Fuzzing process finished. ---")

        os.makedirs(LOG_DIR, exist_ok=True)

        if return_code == 0:
            content_to_write = "success"
            message = f"Fuzzing build command completed successfully. Result saved to '{LOG_FILE_PATH}'."
            status = "success"
        else:
            content_to_write = "".join(log_buffer)
            message = f"Fuzzing build command failed. Detailed log saved to '{LOG_FILE_PATH}'."
            status = "error"

        with open(LOG_FILE_PATH, "w", encoding="utf-8") as f:
            f.write(content_to_write)

        print(message)
        return {"status": status, "message": message}

    except Exception as e:
        message = f"An unknown exception occurred while executing the fuzzing command: {str(e)}"
        print(message)
        os.makedirs(LOG_DIR, exist_ok=True)
        with open(LOG_FILE_PATH, "w", encoding="utf-8") as f:
            f.write(message)
        return {"status": "error", "message": message}
